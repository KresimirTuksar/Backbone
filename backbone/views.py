from .forms import BiljeskaForm, DionicaEditForm, DionicaForm, DocumentForm, IzvrsavanjeForm, KabelEditForm, KabelForm, MjestoForm, OdkForm, PodrucjeForm, KabelSearchForm
from django.http.response import Http404
from django.http import HttpResponse, HttpResponseRedirect
from django.shortcuts import get_object_or_404, redirect, render
import datetime
from .models import *
from django.contrib.auth.decorators import login_required
from django.core.files.storage import FileSystemStorage

import openpyxl
# Create your views here.


@login_required
def mjesta(request):
    mjesto = Mjesto.objects.all()

    context = {'mjesto': mjesto}

    return render(request, 'home.html', context)


@login_required
def mjesto_novo(request):
    mjesto = Mjesto.objects.all()
    form = MjestoForm()
    title = 'Unos mjesta'
    if request.method == 'POST':
        form = MjestoForm(request.POST)
        if form.is_valid():
            instance = form.save(commit=False)
            instance.kreirao = request.user
            instance.save()

            return redirect('mjesta')

    context = {'mjesto': mjesto, 'title': title, 'form': form}
    return render(request, 'forma.html', context)


@login_required
def podrucja(request, pk):
    mjesto = get_object_or_404(Mjesto, pk=pk)

    context = {'mjesto': mjesto}
    return render(request, 'podrucje.html', context)


@login_required
def podrucje_novo(request, pk):
    mjesto = get_object_or_404(Mjesto, pk=pk)
    form = PodrucjeForm()
    title = 'Područje - novi unos'

    if request.method == 'POST':
        form = PodrucjeForm(request.POST)
        if form.is_valid():
            instance = form.save(commit=False)
            instance.mjesto = mjesto
            instance.kreirao = request.user
            instance.save()
            return redirect('podrucja', pk=mjesto.pk)

    context = {'mjesto': mjesto, 'form': form, 'title': title}
    return render(request, 'forma.html', context)


@login_required
def odk_novo(request, pk, podrucje_pk):
    podrucje = get_object_or_404(Podrucje, mjesto__pk=pk, pk=podrucje_pk)
    form = OdkForm()
    title = 'Odk - novi unos'

    if request.method == 'POST':
        form = OdkForm(request.POST)
        if form.is_valid():
            instance = form.save(commit=False)
            instance.podrucje = podrucje
            instance.kreirao = request.user
            instance.save()
            return redirect('odk', podrucje_pk=podrucje.pk, pk=podrucje.mjesto.pk)

    context = {'podrucje': podrucje, 'form': form, 'title': title}
    return render(request, 'forma.html', context)


@login_required
def dionice(request, pk, podrucje_pk):
    podrucje = get_object_or_404(Podrucje, mjesto__pk=pk, pk=podrucje_pk)
    

    context = {'podrucje': podrucje}

    return render(request, 'dionice.html', context)


@login_required
def dionice_novo(request, pk, podrucje_pk):
    podrucje = get_object_or_404(Podrucje, mjesto__pk=pk, pk=podrucje_pk)
    title = 'Dionica - novi unos'
    form = DionicaForm()
    if request.method == 'POST':
        form = DionicaForm(request.POST)
        if form.is_valid():
            instance = form.save(commit=False)
            instance.podrucje = podrucje
            instance.kreirao = request.user
            instance.save()
            return redirect('dionice', podrucje_pk=podrucje.pk, pk=podrucje.mjesto.pk)

    context = {'podrucje': podrucje, 'form': form, 'title': title}
    return render(request, 'forma.html', context)


@login_required
def dionica_uredi(request, pk, podrucje_pk, dionica_pk):
    dionica = get_object_or_404(
        Dionica, podrucje__mjesto__pk=pk, podrucje__pk=podrucje_pk, pk=dionica_pk)
    form = DionicaEditForm(instance=dionica)
    title = 'Dionica - Uredi'

    if request.method == 'POST':
        form = DionicaEditForm(request.POST, instance=dionica)
        if form.is_valid():
            form.save()
            return redirect('dionice', podrucje_pk=dionica.podrucje.pk, pk=dionica.podrucje.mjesto.pk)

    context = {'dionica': dionica, 'form': form, 'title': title}
    return render(request, 'forma.html', context)


@login_required
def kablovi(request, pk, podrucje_pk, dionica_pk):
    dionica = get_object_or_404(
        Dionica, podrucje__mjesto__pk=pk, podrucje__pk=podrucje_pk, pk=dionica_pk)

    context = {'dionica': dionica}

    return render(request, 'kablovi.html', context)

@login_required
def kablovi_odk(request, pk, podrucje_pk, odk_pk):
    odk = get_object_or_404(
        Odk, podrucje__mjesto__pk=pk, podrucje__pk=podrucje_pk, pk=odk_pk)

    context = {'odk': odk}

    return render(request, 'odk.html', context)


@login_required
def kablovi_novo(request, pk, podrucje_pk, dionica_pk):
    dionica = get_object_or_404(
        Dionica, podrucje__mjesto__pk=pk, podrucje__pk=podrucje_pk, pk=dionica_pk)
    form = KabelForm()
    title = 'Kabel - novi unos'

    if request.method == 'POST':
        form = KabelForm(request.POST)
        if form.is_valid():
            instance = form.save(commit=False)
            instance.dionica = dionica
            instance.kreirao = User.objects.first()
            instance.save()
            return redirect('kablovi', dionica_pk=dionica.pk, podrucje_pk=dionica.podrucje.pk, pk=dionica.podrucje.mjesto.pk)

    context = {'dionica': dionica, 'form': form, 'title': title}
    return render(request, 'forma.html', context)


@login_required
def kabel(request, pk, podrucje_pk, dionica_pk, kabel_pk):
    kabel = get_object_or_404(Kabel, dionica__podrucje__mjesto__pk=pk,
                              dionica__podrucje__pk=podrucje_pk, dionica__pk=dionica_pk, pk=kabel_pk)
    context = {'kabel': kabel}

    return render(request, 'detalji.html', context)


@login_required
def kabel_uredi(request, pk, podrucje_pk, dionica_pk, kabel_pk):
    kabel = get_object_or_404(Kabel, dionica__podrucje__mjesto__pk=pk,
                              dionica__podrucje__pk=podrucje_pk, dionica__pk=dionica_pk, pk=kabel_pk)
    form = KabelEditForm(instance=kabel)
    title = 'Uredi kabel'

    if request.method == 'POST':
        form = KabelEditForm(request.POST, instance=kabel)
        if form.is_valid():
            instance = form.save(commit=False)
            instance.radna_duzina = abs(instance.pocetna - instance.zavrsna)
            if (instance.status != 'Riješeno'):
                instance.rijeseno = False
            instance.save()
            next = request.POST.get('next', '/')
            return HttpResponseRedirect(next)
            
            #return redirect('kablovi', dionica_pk=kabel.dionica.pk, podrucje_pk=kabel.dionica.podrucje.pk, pk=kabel.dionica.podrucje.mjesto.pk)

    context = {'kabel': kabel, 'form': form, 'title': title}
    return render(request, 'forma.html', context)


@login_required
def kabel_odradi(request, pk, podrucje_pk, dionica_pk, kabel_pk):
    kabel = get_object_or_404(Kabel, dionica__podrucje__mjesto__pk=pk,
                              dionica__podrucje__pk=podrucje_pk, dionica__pk=dionica_pk, pk=kabel_pk)
    form = IzvrsavanjeForm(instance=kabel)
    title = 'Radni unos'
    if(kabel.rijeseno == False):
        
        if request.method == 'POST':
            form = IzvrsavanjeForm(request.POST, instance=kabel)
            if form.is_valid():
                instance = form.save(commit=False)
                
                instance.radna_duzina = abs(instance.pocetna - instance.zavrsna)
                instance.zavrseno_datum = datetime.datetime.now()
                instance.zavrsio = User.objects.first()
                if(instance.status == 'Riješeno'):
                    instance.rijeseno = True

                instance.save()

                next = request.POST.get('next', '/')
                return HttpResponseRedirect(next)
                #return redirect('kablovi', dionica_pk=kabel.dionica.pk, podrucje_pk=kabel.dionica.podrucje.pk, pk=kabel.dionica.podrucje.mjesto.pk)
    else:
        
        return redirect('kablovi', dionica_pk=kabel.dionica.pk, podrucje_pk=kabel.dionica.podrucje.pk, pk=kabel.dionica.podrucje.mjesto.pk)

    context = {'kabel': kabel, 'form': form, 'title': title}
    return render(request, 'forma.html', context)




@login_required
def biljeska_novo(request, pk, podrucje_pk, dionica_pk, kabel_pk):
    kabel = get_object_or_404(Kabel, dionica__podrucje__mjesto__pk=pk,
                              dionica__podrucje__pk=podrucje_pk, dionica__pk=dionica_pk, pk=kabel_pk)
    form = BiljeskaForm()
    title = 'Bilješka - novi unos'

    if request.method == 'POST':
        form = BiljeskaForm(request.POST)
        if form.is_valid():
            instance = form.save(commit=False)
            instance.kabel = kabel
            instance.kreirao = request.user
            instance.save()
            return redirect('detalji', kabel_pk=kabel.pk, dionica_pk=kabel.dionica.pk, podrucje_pk=kabel.dionica.podrucje.pk, pk=kabel.dionica.podrucje.mjesto.pk)

    context = {'kabel': kabel, 'form': form, 'title': title}
    return render(request, 'forma.html', context)

#za ovo vjerojatno treba podrška
@login_required
def model_form_upload(request, pk, podrucje_pk):
    podrucje = get_object_or_404(Podrucje, mjesto__pk=pk, pk=podrucje_pk)
    if request.method == 'POST':
        form = DocumentForm(request.POST, request.FILES)
        if form.is_valid():
            form.save()
            handle_file = handle_parameters_upload(
                request, request.FILES['document'], pk, podrucje_pk)
            return redirect('dionice', podrucje_pk=podrucje.pk, pk=podrucje.mjesto.pk)
    else:
        form = DocumentForm()

    context={'form': form, 'podrucje':podrucje}

    return render(request, 'upload.html', context)


def handle_parameters_upload(request, file, pk, podrucje_pk):
    podrucje = get_object_or_404(Podrucje, mjesto__pk=pk, pk=podrucje_pk)
    wb = openpyxl.load_workbook(file, read_only=True)
    ws = wb.active

    if podrucje.je_uvezeno == False: #provjera jeli područje več kreirano
        for row in ws.iter_rows(min_row=2):
            if row[3].value is not None: #provjera jeli red prazan

                dionica = Dionica(
                    podrucje=podrucje,
                    naziv=row[3].value,
                    rb=row[0].value,
                    kreirao=User.objects.first(),
                    duzina=row[4].value
                )
                dionica.save()

                kablovi_na_dionici = row[10].value.splitlines()
                for x in kablovi_na_dionici:
                    x = x.strip()
                    
                    odk, created = Odk.objects.get_or_create(
                        naziv=x[:6]+'_'+podrucje.naziv,
                        podrucje=podrucje,
                        kreirao=User.objects.first(),
                        defaults={},
                    )

                    kabel = Kabel(
                        segment=x[:10],
                        dionica=dionica,
                        odk=odk,
                        # broj_niti=x[11],
                        napomena=x,
                        kreirao=User.objects.first()
                    )
                    kabel.save()

                    podrucje.je_uvezeno=True
                    podrucje.save()
            else:
                break
    
    else:
        pass
        #dovršiti update iz excela
        """ for row in ws.iter_rows(min_row=2):
            if row[3].value is not None: #provjera jeli red prazan


                dionica, created = Dionica.objects.get_or_create(
                        naziv=row[3].value,
                        podrucje=podrucje,
                        rb=row[0].value,
                        kreirao=User.objects.first(),
                        defaults={},
                    )

                kablovi_na_dionici = row[10].value.splitlines()
                for x in kablovi_na_dionici:
                    x = x.strip()
                    
                    odk, created = Odk.objects.get_or_create(
                        naziv=x[:6]+'_'+podrucje.naziv,
                        podrucje=podrucje,
                        kreirao=User.objects.first(),
                        defaults={},
                    )

                    kabel, created = Kabel.objects.get_or_create(
                        segment=x[:10],
                        dionica=dionica,
                        napomena=x,
                        kreirao=User.objects.first(),
                        defaults={},
                    )

            else:
                break """

